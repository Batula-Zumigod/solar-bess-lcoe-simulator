import numpy as np
import pandas as pd
import requests
from datetime import datetime, timedelta
from io import BytesIO
from tqdm import tqdm
import warnings
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import pvlib
from pvlib import location, irradiance, atmosphere, pvsystem


class WeatherDataDownloader:
    """Download solar and weather data from PVGIS"""
    
    @staticmethod
    def download_tmy_pvgis(latitude, longitude):
        """
        Download Typical Meteorological Year (TMY) data from PVGIS.
        Returns DataFrame with columns: datetime, ghi, dni, dhi, temp_air, wind_speed
        """
        try:
            print(f"🌤️ Downloading TMY data for location...")
            print(f"    Latitude: {latitude}, Longitude: {longitude}")
        
            # PVGIS API endpoint
            url = "https://re.jrc.ec.europa.eu/api/tmy"
        
            params = {
                'lat': latitude,
                'lon': longitude,
                'outputformat': 'csv',
            }
        
            print("    Connecting to PVGIS API...")
            response = requests.get(url, params=params, timeout=30)
        
            if response.status_code != 200:
                raise ConnectionError(f"PVGIS API returned status code: {response.status_code}")
        
            if len(response.content) < 100:
                raise ConnectionError("Received empty or invalid response from PVGIS")
        
            print("    Parsing CSV data...")
        
            # Parse response content directly
            from io import StringIO
            lines = response.text.split('\n')
        
            # Find the line with column headers (contains "time(UTC)")
            header_line = 0
            for i, line in enumerate(lines):
                if 'time(UTC)' in line:
                    header_line = i
                    break
        
            # Skip lines before header
            csv_data = '\n'.join(lines[header_line:])
            tmy_df = pd.read_csv(StringIO(csv_data))
        
            # Clean column names
            tmy_df.columns = tmy_df.columns.str.strip()
        
            # Filter valid data rows
            if 'time(UTC)' in tmy_df.columns:
                # Keep only rows matching datetime format (YYYYMMDD:HHMM)
                tmy_df = tmy_df[tmy_df['time(UTC)'].astype(str).str.match(r'^\d{8}:\d{4}$', na=False)]
            
                # Parse datetime in UTC
                tmy_df['datetime'] = pd.to_datetime(tmy_df['time(UTC)'], format='%Y%m%d:%H%M', errors='coerce')
            
                # Convert from UTC to Ulaanbaatar timezone
                tmy_df['datetime'] = tmy_df['datetime'].dt.tz_localize('UTC').dt.tz_convert('Asia/Ulaanbaatar')
            
                print("    ✓ Converted timezone from UTC to Asia/Ulaanbaatar (UTC+8)")
            else:
                raise ValueError("Cannot find 'time(UTC)' column in CSV")
        
            # Drop rows with invalid datetime
            tmy_df = tmy_df.dropna(subset=['datetime'])
        
            # Map PVGIS column names to standard names
            column_mapping = {
                'T2m': 'temp_air',
                'G(h)': 'ghi',
                'Gb(n)': 'dni',
                'Gd(h)': 'dhi',
                'WS10m': 'wind_speed',
                'RH': 'relative_humidity',
                'SP': 'pressure'
            }
        
            tmy_df = tmy_df.rename(columns=column_mapping)
        
            # Sort by datetime
            tmy_df = tmy_df.sort_values('datetime').reset_index(drop=True)
        
            # Reindex to start from midnight (00:00) in Ulaanbaatar time
            tmy_df['hour'] = tmy_df['datetime'].dt.hour
            tmy_df['day_of_year'] = tmy_df['datetime'].dt.dayofyear
            tmy_df = tmy_df.sort_values(['day_of_year', 'hour'])
            tmy_df = tmy_df.drop(['hour', 'day_of_year'], axis=1)
        
            # Reset to sequential hourly index starting from Jan 1 00:00
            start_date = pd.Timestamp('2002-01-01 00:00:00', tz='Asia/Ulaanbaatar')
            new_index = pd.date_range(start=start_date, periods=len(tmy_df), freq='h')
            tmy_df['datetime'] = new_index
        
            print("    ✓ Reorganized data to start at midnight (00:00) Ulaanbaatar time")
        
            # Keep only needed columns
            needed_cols = ['datetime', 'ghi', 'dni', 'dhi', 'temp_air', 'wind_speed']
            available_cols = [col for col in needed_cols if col in tmy_df.columns]
            tmy_df = tmy_df[available_cols]
        
            # Convert to numeric
            for col in tmy_df.columns:
                tmy_df[col] = pd.to_numeric(tmy_df[col], errors='coerce')
        
    
        
            print(f"  ✓ Downloaded {len(tmy_df)} hours of TMY data")
            if 'temp_air' in tmy_df.columns:
                print(f"    Average temperature: {tmy_df['temp_air'].mean():.1f}°C")
            if 'ghi' in tmy_df.columns:
                print(f"    Average GHI: {tmy_df['ghi'].mean():.0f} W/m²")
        
            # Validate required columns
            required_cols = ['datetime', 'ghi', 'dhi', 'dni', 'temp_air']
            missing_cols = [col for col in required_cols if col not in tmy_df.columns]

            if missing_cols:
                raise ValueError(f"Weather data missing required columns: {missing_cols}")

            # Validate no NaN in critical columns
            if tmy_df[required_cols].isnull().any().any():
                raise ValueError("Weather data contains NaN values in critical columns")
            tmy_df.to_csv("tmy_ub_pvgis.csv", index=False)
            return tmy_df
        
        except requests.exceptions.Timeout:
            raise ConnectionError("Request timed out. Please check your internet connection.")
        except requests.exceptions.ConnectionError:
            raise ConnectionError("Cannot connect to PVGIS server. Please check your internet connection.")
        except Exception as e:
            raise ConnectionError(f"Error downloading TMY data: {str(e)}")    


class PVSystem:
    """PV system modeling with PVLib for accurate fixed-tilt calculations"""
    
    def __init__(self, capacity_kwp, PV_System_Losses, pv_lid, pv_degradation,
                 temp_coeff, inverter_efficiency, acdc_ratio,
                 tilt, azimuth, latitude, longitude, timezone, TimeStep):
        """
        Parameters (ALL REQUIRED - no defaults):
        -----------
        capacity_kwp : float - PV system capacity in kWp
        temp_coeff : float - Temperature coefficient (%/°C)
        PV_System_Losses : float - Total system losses (soiling, wiring, mismatch, etc.)
        inverter_efficiency : float - Inverter efficiency
        tilt : float - Panel tilt angle in degrees
        azimuth : float - Panel azimuth in degrees (180=south, 90=east, 270=west)
        latitude : float - Site latitude for solar position calculations
        longitude : float - Site longitude for solar position calculations
        timezone : str - Timezone string (e.g. 'Asia/Ulaanbaatar')
        pv_lid : float - Light-induced degradation factor (first year)
        pv_degradation : float - Annual degradation rate
        """
        # Validate all required parameters
        _required = {
            'capacity_kwp': capacity_kwp, 'PV_System_Losses': PV_System_Losses,
            'pv_lid': pv_lid, 'pv_degradation': pv_degradation,
            'temp_coeff': temp_coeff,
            'inverter_efficiency': inverter_efficiency, 'acdc_ratio': acdc_ratio, 'tilt': tilt,
            'azimuth': azimuth, 'latitude': latitude,
            'longitude': longitude, 'timezone': timezone, 'TimeStep': TimeStep,
        }
        _missing = [k for k, v in _required.items() if v is None]
        if _missing:
            raise ValueError(f"PVSystem: missing required parameters: {_missing}")

        self.capacity_kwp = capacity_kwp
        self.temp_coeff = temp_coeff
        self.PV_System_Losses = PV_System_Losses
        self.inverter_efficiency = inverter_efficiency
        self.acdc_ratio = acdc_ratio
        self.pv_lid = pv_lid
        self.pv_degradation = pv_degradation
        self.latitude = latitude
        self.longitude = longitude
        self.tilt = tilt
        self.azimuth = azimuth
        self.timezone = timezone
        self.TimeStep=TimeStep
    
    def calculate_pv_generation(self, weather_data, year, TimeStep):
        """
        Calculate hourly PV generation using POA irradiance with PVLib
        
        Parameters (ALL REQUIRED - no defaults):
        -----------
        weather_data : pd.DataFrame with ghi, dni, dhi, temp_air, wind_speed, datetime
        year : int - Year of operation (for degradation)
        inverter_capacity_kw : float or None - Inverter AC capacity for clipping (pass None explicitly if unused)
        dt : float - Timestep duration in hours
        
        Returns:
        --------
        np.array of energy generation in kWh per timestep
        """
        if year is None:
            raise ValueError("PVSystem.calculate_pv_generation: 'year' is required and cannot be None")
        if TimeStep is None:
            raise ValueError("PVSystem.calculate_pv_generation: 'dt' is required and cannot be None")
        if self.latitude is None or self.longitude is None:
            raise ValueError("Latitude and longitude required for PV calculations")
        
        # Ensure weather_data is exactly 8760 hours
        if len(weather_data) > 8760:
            weather_data = weather_data.iloc[:8760].copy()
        
        n_hours = len(weather_data)
        pv_output = np.zeros(n_hours)
        
        # Validate all required weather columns exist
        _required_weather = ['ghi', 'dni', 'dhi', 'temp_air', 'wind_speed']
        _missing_weather = [col for col in _required_weather if col not in weather_data.columns]
        if _missing_weather:
            raise ValueError(f"PVSystem.calculate_pv_generation: weather_data missing required columns: {_missing_weather}")

        # Extract weather arrays
        ghi = weather_data['ghi'].values
        dni = weather_data['dni'].values
        dhi = weather_data['dhi'].values
        temp_air = weather_data['temp_air'].values
        wind_speed = weather_data['wind_speed'].values
        
        # Create location object
        site_location = location.Location(self.latitude, self.longitude, tz=self.timezone)
        
        # Get datetime index — required, no fallback
        if 'datetime' not in weather_data.columns:
            raise ValueError("PVSystem.calculate_pv_generation: weather_data missing required 'datetime' column")
        times = pd.DatetimeIndex(weather_data['datetime'])
        
        # Calculate solar position for all hours at once
        try:
            solar_position = site_location.get_solarposition(times)
            solar_zenith = solar_position['apparent_zenith'].values
            solar_azimuth = solar_position['azimuth'].values
            
            pvlib_available = True
        except Exception as e:
            print(f"Warning: PVLib solar position failed: {e}")
            pvlib_available = False
            solar_zenith = np.full(n_hours, 45.0)
            solar_azimuth = np.full(n_hours, 180.0)
        
        # Calculate POA irradiance for all hours
        poa_global = np.zeros(n_hours)
        
        for i in range(n_hours):
            if ghi[i] > 10 and pvlib_available:  # Only calculate if there's meaningful irradiance
                try:
                    poa_irradiance = irradiance.get_total_irradiance(
                        surface_tilt=self.tilt,
                        surface_azimuth=self.azimuth,
                        solar_zenith=solar_zenith[i],
                        solar_azimuth=solar_azimuth[i],
                        dni=dni[i],
                        ghi=ghi[i],
                        dhi=dhi[i],
                        albedo=0.2
                    )
                    poa_global[i] = poa_irradiance['poa_global']
                except Exception:
                    # Fallback to simple cosine model
                    poa_global[i] = ghi[i] * max(0, np.cos(np.radians(solar_zenith[i])))
            elif ghi[i] > 10:
                # Fallback calculation without PVLib
                poa_global[i] = ghi[i] * max(0, np.cos(np.radians(self.tilt)))
            else:
                poa_global[i] = 0
        
        # Vectorized temperature and power calculations
        # Wind speed at 2m height (for cell temperature)
        z0 = 0.03
        wind_speed_2m = wind_speed * (np.log(2 / z0) / np.log(10 / z0))
        
        U0 = 25.0
        U1 = 6.84
        U = U0 + U1 * wind_speed_2m
        absorbed_irradiance = poa_global * 0.9  # 90% absorption
        cell_temp = temp_air + (absorbed_irradiance / U) * (1 - 0.20)  # For the time being lets call our module is 20% efficient
        
        # Temperature derating
        temp_loss = 1 + self.temp_coeff * (cell_temp - 25) / 100  # temp_coeff is in %/°C
        
        # DC power calculation
        dc_power = self.capacity_kwp * (poa_global / 1000) * temp_loss
        
        # Apply system losses
        dc_power = dc_power * (1 - self.PV_System_Losses)
        
        # Apply LID (first year only) and degradation
        if year == 0:
            degradation_factor = (1 - self.pv_lid)
        else:
            degradation_factor = (1 - self.pv_lid) * (1 - self.pv_degradation) ** year
        
        dc_power = dc_power * degradation_factor
        
        # Convert to DC to AC with inverter efficiency
        ac_power = dc_power * self.inverter_efficiency
        
        # Apply inverter clipping if specified
        if self.acdc_ratio is not None:
            ac_power = np.minimum(ac_power, self.capacity_kwp / self.acdc_ratio)
        
        # Convert power to energy (kWh)
        pv_output = np.maximum(0, ac_power) * self.TimeStep
        
        return pv_output


class BatterySystem:
    """Battery Energy Storage System with degradation modeling"""
    
    def __init__(self, capacity_kwh, bess_hours, bessefficiency,
                 soc_min, soc_max, TimeStep):

        # Validate all required parameters
        _required = {
            'capacity_kwh': capacity_kwh, 'bess_hours': bess_hours,
            'bessefficiency': bessefficiency, 'soc_min': soc_min,
            'soc_max': soc_max,'TimeStep': TimeStep,
        }
        _missing = [k for k, v in _required.items() if v is None]
        if _missing:
            raise ValueError(f"BatterySystem: missing required parameters: {_missing}")
        self.nominal_capacity_kwh = capacity_kwh
        self.capacity_kwh = capacity_kwh
        self.bess_hours=bess_hours
        self.power_kw = self.capacity_kwh / bess_hours
        self.efficiency = bessefficiency
        self.soc_min = soc_min
        self.soc_max = soc_max
        self.soc = 0.5
        self.TimeStep=TimeStep
        
        self.cumulative_throughput_kwh = 0
        self.replacement_count = 0
        self.replacement_years = []
    
    def get_capacity_fade(self, years_elapsed):
        """Calculate total capacity fade from calendar and cycle aging"""
        days_elapsed = years_elapsed * 365.25

        k_cal = 1.754e-3
        k_cyc = 1.755e-3
        
        calendar_fade = k_cal * np.sqrt(days_elapsed)

        efc = self.cumulative_throughput_kwh / (2 * self.nominal_capacity_kwh)
        cycle_fade = k_cyc * np.sqrt(efc)
        
        total_fade = 1 - (1 - calendar_fade) * (1 - cycle_fade)
        
        return min(total_fade, 1.0)
    
    def update_capacity(self, years_elapsed):
        """Update battery capacity based on degradation"""
        fade = self.get_capacity_fade(years_elapsed)
        self.capacity_kwh = self.nominal_capacity_kwh * (1 - fade)
        self.power_kw = self.capacity_kwh / self.bess_hours
        
        return 1 - fade
    
    def replace_battery(self, current_year):
        """Replace battery - reset to new condition"""
        self.capacity_kwh = self.nominal_capacity_kwh
        self.power_kw = self.nominal_capacity_kwh / self.bess_hours
        self.cumulative_throughput_kwh = 0
        self.total_cycles = 0
        self.soc = 0.5
        self.replacement_count += 1
        self.replacement_years.append(current_year)
    
    def charge(self, power_kw):
        """Charge battery with given power for dt hours. dt is required — no default."""
        if self.TimeStep is None:
            raise ValueError("BatterySystem.charge: 'dt' is required and cannot be None")
        
        available_storage_kwh = (self.soc_max - self.soc) * self.capacity_kwh
        charge_efficiency = np.sqrt(self.efficiency)
        required_input_kwh = available_storage_kwh / charge_efficiency
        
        max_charge_power_from_capacity = required_input_kwh / self.TimeStep
        
        max_charge_power = min(self.power_kw, max_charge_power_from_capacity)
        
        actual_charge_power = min(power_kw, max_charge_power)
        
        energy_from_pv = actual_charge_power * self.TimeStep  # Energy taken from PV
        energy_stored = energy_from_pv * charge_efficiency  # Energy stored in battery
        
        self.soc += energy_stored / self.capacity_kwh
        self.soc = min(self.soc, self.soc_max)  # Simple clamp
        
        self.cumulative_throughput_kwh += energy_stored
        
        return energy_from_pv
    
    def discharge(self, power_kw):
        """Discharge battery with given power for dt hours. dt is required — no default."""
        if self.TimeStep is None:
            raise ValueError("BatterySystem.discharge: 'dt' is required and cannot be None")
        
        available_energy_kwh = (self.soc - self.soc_min) * self.capacity_kwh

        discharge_efficiency = np.sqrt(self.efficiency)
        
        max_deliverable_kwh = available_energy_kwh * discharge_efficiency
        
        max_discharge_power_from_capacity = max_deliverable_kwh / self.TimeStep
        
        max_discharge_power = min(self.power_kw, max_discharge_power_from_capacity)
        
        actual_discharge_power = min(power_kw, max_discharge_power)
        
        energy_delivered_pv = actual_discharge_power * self.TimeStep 
        energy_from_battery = energy_delivered_pv / discharge_efficiency
        
        self.soc -= energy_from_battery / self.capacity_kwh
        self.soc = max(self.soc, self.soc_min) 
        
        self.cumulative_throughput_kwh += energy_from_battery
        
        return energy_delivered_pv


class HybridEnergySimulator:
    """Simulator for PV+BESS+Backup systems with grid limit support and unmet load tracking"""
    
    def __init__(self, pv_system, bess_system,
                 load_profile, weather_data,
                 backup_type, diesel_fuel_consumption,
                 diesel_fuel_price, noon_grid_tariff, peak_grid_tariff, night_grid_tariff,
                 grid_limit_kw,TimeStep):

        # Validate required parameters
        _required = {
            'load_profile': load_profile, 'weather_data': weather_data,
            'backup_type': backup_type, 'diesel_fuel_consumption': diesel_fuel_consumption,
            'diesel_fuel_price': diesel_fuel_price, 'noon_grid_tariff': noon_grid_tariff, 
            'peak_grid_tariff': peak_grid_tariff, 'night_grid_tariff': night_grid_tariff,
            'grid_limit_kw': grid_limit_kw, 'TimeStep': TimeStep
        }
        _missing = [k for k, v in _required.items() if v is None]
        if _missing:
            raise ValueError(f"HybridEnergySimulator: missing required parameters: {_missing}")
        if 'Load_kW' not in load_profile.columns:
            raise ValueError("HybridEnergySimulator: load_profile must contain 'Load_kW' column")
        if len(load_profile) < 8760:
            raise ValueError(f"HybridEnergySimulator: load_profile has {len(load_profile)} rows, need 8760")

        self.pv = pv_system
        self.bess = bess_system
        self.load_profile = load_profile
        self.weather_data = weather_data
        self.backup_type = backup_type
        self.diesel_fuel_consumption = diesel_fuel_consumption
        self.diesel_fuel_price = diesel_fuel_price
        self.noon_grid_tariff = noon_grid_tariff
        self.peak_grid_tariff = peak_grid_tariff
        self.night_grid_tariff = night_grid_tariff
        self.grid_limit_kw = grid_limit_kw
        self.TimeStep=TimeStep
        
    def simulate_year(self, year):
        
        # ====================================================================
        # 1. CALCULATE PV GENERATION
        # ====================================================================
        pv_generation = np.zeros(8760)
        if self.pv is not None and self.weather_data is not None:
            pv_generation = self.pv.calculate_pv_generation(
                self.weather_data, year, self.TimeStep
            )
        
        # ====================================================================
        # 2. CREATE GRID TARIFF PROFILE (if using grid)
        # ====================================================================
        if self.backup_type in ['Grid', 'Grid+Diesel']:
            grid_tariff_profile = np.array([
                self.get_grid_tariff(h % 24, self.noon_grid_tariff, 
                                    self.peak_grid_tariff, self.night_grid_tariff) 
                for h in range(8760)
            ])
        else:
            grid_tariff_profile = np.full(8760, 0)
        
        # ====================================================================
        # 3. INITIALIZE VARIABLES
        # ====================================================================
        load = self.load_profile['Load_kW'].values[:8760]
        
        # Energy flow tracking arrays
        pv_to_load = np.zeros(8760)
        pv_to_battery = np.zeros(8760)
        battery_to_load = np.zeros(8760)
        generator_to_battery = np.zeros(8760)
        grid_to_battery = np.zeros(8760)
        grid_to_load = np.zeros(8760)
        diesel_to_load = np.zeros(8760)
        unmet_load = np.zeros(8760)
        curtailed = np.zeros(8760)
        generator_spillage = np.zeros(8760)
        soc_profile = []
        
        Generator_max = load.max()
        
        # ====================================================================
        # 4. HOURLY SIMULATION LOOP
        # ====================================================================
        for i in range(8760):
            pv_avail = pv_generation[i]
            load_demand = load[i]
            current_hour = i % 24
            current_tariff = grid_tariff_profile[i]
            
            # Track initial SOC
            if self.bess is not None:
                current_soc = self.bess.soc
                soc_profile.append(current_soc)
            else:
                current_soc = 0
                soc_profile.append(0)
            
            # ================================================================
            # CASE 1: PV GENERATION EXCEEDS LOAD
            # ================================================================
            if pv_avail >= load_demand:
                # Use PV to meet load
                pv_to_load[i] = load_demand
                
                # Calculate excess
                excess_pv = pv_avail - load_demand
                
                # Charge battery with FREE PV excess
                if self.bess is not None and excess_pv > 0:
                    charged = self.bess.charge(excess_pv)
                    pv_to_battery[i] = charged
                    curtailed[i] = excess_pv - charged
                else:
                    curtailed[i] = excess_pv
            
            # ================================================================
            # CASE 2: LOAD EXCEEDS PV GENERATION
            # ================================================================
            else:
                # Step 1: Use all available PV for load
                pv_to_load[i] = pv_avail
                deficit = load_demand - pv_avail
                
                # ============================================================
                # DIESEL GENERATOR DISPATCH
                # ============================================================
                if self.backup_type == 'Diesel':
                    
                    if self.bess is not None and deficit > 0:
                        deliverable = min(deficit, self.bess.power_kw *self.TimeStep, (current_soc - self.bess.soc_min) * self.bess.capacity_kwh * np.sqrt(self.bess.efficiency))
                        if deliverable == deficit:
                            discharged = self.bess.discharge(deficit)
                            battery_to_load[i] = discharged
                            deficit = 0
                    
                    # Use diesel if BESS is not enough for deficit.
                        else:
                            if deficit < 0.7 * Generator_max:
                                generator_output = 0.7 * Generator_max
                                diesel_to_load[i] = deficit
                                excess_generator = generator_output - deficit
                                deficit = 0
                                
                                if self.bess is not None and excess_generator > 0:
                                    charged_from_generator = self.bess.charge(excess_generator)
                                    generator_to_battery[i] = charged_from_generator
                                    generator_spillage[i] = excess_generator - charged_from_generator
                            else:
                                diesel_to_load[i] = deficit
                                deficit = 0
                        
                        # Check if load was fully met
                        if deficit > 1:
                            unmet_load[i] = deficit
                                
                # ============================================================
                # GRID-ONLY DISPATCH WITH TOU AND LIMIT
                # ============================================================
                elif self.backup_type == 'Grid':

                    
                    # Determine how much grid is available
                    grid_available = self.grid_limit_kw
                    
                    if current_tariff <= self.night_grid_tariff:
                        # OFF-PEAK: Use cheap grid first, then battery if grid limit is reached
                        grid_used = min(deficit, grid_available)
                        grid_to_load[i] = grid_used
                        deficit -= grid_used
                        
                        # If grid limit is reached and there's still deficit, use battery
                        if deficit > 0 and self.bess is not None and current_soc > self.bess.soc_min:
                            discharged = self.bess.discharge(deficit)
                            battery_to_load[i] = discharged
                            deficit -= discharged
                        
                        # Charge battery from cheap grid if capacity available and no deficit
                        if (deficit == 0 and
                            self.bess is not None and 
                            current_soc < self.bess.soc_max and 
                            grid_used < grid_available and
                            self.should_charge_from_grid(
                                current_hour, current_tariff, i, 
                                pv_generation, current_soc
                            )):
                            
                            remaining_grid = grid_available - grid_used
                            max_charge_power = min(self.bess.power_kw, remaining_grid)
                            available_capacity = (self.bess.soc_max - current_soc) * self.bess.capacity_kwh
                            charge_amount = min(max_charge_power, available_capacity)
                            
                            charged = self.bess.charge(charge_amount)
                            grid_to_battery[i] = charged
                    
                    elif current_tariff >= self.peak_grid_tariff:
                        # PEAK: Use battery first, then grid, combining both if needed
                        if self.bess is not None and deficit > 0 and current_soc > self.bess.soc_min:
                            discharged = self.bess.discharge(deficit)
                            battery_to_load[i] = discharged
                            deficit -= discharged
                        
                        # Use remaining grid capacity to cover any remaining deficit
                        if deficit > 0:
                            grid_used = min(deficit, grid_available)
                            grid_to_load[i] = grid_used
                            deficit -= grid_used
                    
                    else:
                        # MID-PEAK: Smart strategy - use grid fully first, then battery for remainder
                        # This ensures we utilize available grid capacity before draining battery
                        if deficit <= grid_available:
                            # If grid can cover entire load, decide based on SOC
                            if self.bess is not None and current_soc > 0.9*self.bess.soc_max:
                                # High SOC: Use battery first to save grid capacity
                                available_above_90 = (current_soc - 0.9*self.bess.soc_max) * self.bess.capacity_kwh * np.sqrt(self.bess.efficiency)
                                discharge_amount = min(deficit, available_above_90)
                                discharged = self.bess.discharge(discharge_amount)
                                battery_to_load[i] += discharged
                                deficit -= discharged
                                current_soc = self.bess.soc  # update local soc tracker
                                
                                # Use grid for any remainder                            
                                if deficit > 0:
                                    grid_used = min(deficit, grid_available)
                                    grid_to_load[i] = grid_used
                                    deficit -= grid_used
                
                            else:
                                # Low/moderate SOC: Use grid to preserve battery
                                grid_used = min(deficit, grid_available)
                                grid_to_load[i] = grid_used
                                deficit -= grid_used
                        else:
                            # Grid cannot cover entire load: use grid fully, then battery
                            grid_used = min(deficit, grid_available)
                            grid_to_load[i] = grid_used
                            deficit -= grid_used
                            
                            # Use battery for remainder
                            if deficit > 0 and self.bess is not None and current_soc > self.bess.soc_min:
                                discharged = self.bess.discharge(deficit)
                                battery_to_load[i] = discharged
                                deficit -= discharged
                    
                    # Track unmet load if grid limit reached and battery exhausted
                    if deficit > 1e-6:
                        unmet_load[i] = deficit

                
                # ============================================================
                # GRID+DIESEL HYBRID DISPATCH WITH GRID LIMIT
                # PRIORITY varies by tariff period, diesel always last resort
                # ============================================================
                elif self.backup_type == 'Grid+Diesel':

                    grid_available = self.grid_limit_kw  # 0 means no grid

                    # ----------------------------------------------------------
                    # HELPER: Diesel backup — runs after grid+bess have tried
                    # ----------------------------------------------------------
                    def _dispatch_diesel(deficit):
                        if deficit <= 0:
                            return deficit
                        if deficit < 0.7 * Generator_max:
                            # Generator must run at minimum 70% load
                            gen_output = 0.7 * Generator_max
                            diesel_to_load[i] = deficit
                            excess = gen_output - deficit
                            deficit = 0
                            if self.bess is not None and excess > 0:
                                charged = self.bess.charge(excess)
                                generator_to_battery[i] = charged
                                generator_spillage[i] += excess - charged
                        else:
                            diesel_to_load[i] = deficit
                            deficit = 0
                        return deficit

                    # ----------------------------------------------------------
                    # OFF-PEAK / NIGHT  (cheapest tariff)
                    # Grid up to limit → BESS for remainder → Diesel last resort
                    # If no deficit and grid headroom: opportunistically charge BESS
                    # ----------------------------------------------------------
                    if current_tariff <= self.night_grid_tariff:

                        # 1. Grid first
                        grid_used = min(deficit, grid_available)
                        grid_to_load[i] = grid_used
                        deficit -= grid_used

                        # 2. BESS covers what grid couldn't
                        if deficit > 0 and self.bess is not None and current_soc > self.bess.soc_min:
                            discharged = self.bess.discharge(deficit)
                            battery_to_load[i] = discharged
                            deficit -= discharged

                        # 3. Diesel last resort
                        deficit = _dispatch_diesel(deficit)

                        # 4. Opportunistic grid → BESS charging (only if load is fully met)
                        if (deficit == 0
                                and self.bess is not None
                                and current_soc < self.bess.soc_max
                                and grid_used < grid_available
                                and self.should_charge_from_grid(
                                    current_hour, current_tariff, i, pv_generation, current_soc)):
                            remaining_grid = grid_available - grid_used
                            charge_headroom = (self.bess.soc_max - current_soc) * self.bess.capacity_kwh
                            charge_amount = min(self.bess.power_kw, remaining_grid, charge_headroom)
                            charged = self.bess.charge(charge_amount)
                            grid_to_battery[i] = charged

                    # ----------------------------------------------------------
                    # PEAK  (most expensive tariff)
                    # BESS first → Grid for remainder → Diesel last resort
                    # ----------------------------------------------------------
                    elif current_tariff >= self.peak_grid_tariff:

                        # 1. BESS first
                        if self.bess is not None and deficit > 0 and current_soc > self.bess.soc_min:
                            discharged = self.bess.discharge(deficit)
                            battery_to_load[i] = discharged
                            deficit -= discharged

                        # 2. Grid covers what BESS couldn't
                        if deficit > 0:
                            grid_used = min(deficit, grid_available)
                            grid_to_load[i] = grid_used
                            deficit -= grid_used

                        # 3. Diesel last resort
                        deficit = _dispatch_diesel(deficit)

                    # ----------------------------------------------------------
                    # NOON / MID-PEAK  (average tariff, peak hours approaching)
                    # Conservative on BESS: only use first if SOC > 90%
                    # → SOC > 90%:  BESS first, then Grid
                    # → SOC ≤ 90%:  Grid first, then BESS
                    # Diesel last resort in both cases
                    # ----------------------------------------------------------
                    else:
                        if self.bess is not None and current_soc > 0.9*self.bess.soc_max:
                            # High SOC — discharge BESS first, preserve grid capacity
                            available_above_90 = (current_soc - 0.9*self.bess.soc_max) * self.bess.capacity_kwh * np.sqrt(self.bess.efficiency)
                            discharge_amount = min(deficit, available_above_90)
                            discharged = self.bess.discharge(discharge_amount)
                            battery_to_load[i] += discharged
                            deficit -= discharged
                            current_soc = self.bess.soc

                            if deficit > 0:
                                grid_used = min(deficit, grid_available)
                                grid_to_load[i] = grid_used
                                deficit -= grid_used
                            if deficit > 0 and self.bess is not None and current_soc > self.bess.soc_min:
                                discharged = self.bess.discharge(deficit)
                                battery_to_load[i] += discharged
                                deficit -= discharged

                        else:
                            # Low/moderate SOC — use grid first, preserve BESS for peak
                            grid_used = min(deficit, grid_available)
                            grid_to_load[i] = grid_used
                            deficit -= grid_used

                            if deficit > 0 and self.bess is not None and current_soc > self.bess.soc_min:
                                discharged = self.bess.discharge(deficit)
                                battery_to_load[i] = discharged
                                deficit -= discharged

                        # Diesel last resort
                        deficit = _dispatch_diesel(deficit)

                    # ----------------------------------------------------------
                    # Unmet load — only if ALL sources exhausted
                    # ----------------------------------------------------------
                    if deficit > 1e-6:
                        unmet_load[i] = deficit
                
                # ============================================================
                # NO BACKUP CASE
                # ============================================================
                else:
                    if self.bess is not None and deficit > 0:
                        discharged = self.bess.discharge(deficit)
                        battery_to_load[i] = discharged
                        deficit -= discharged
                    
                    # Remaining deficit is unmet
                    if deficit > 1e-6:
                        unmet_load[i] = deficit
        
        # ====================================================================
        # 5. PREPARE ANNUAL RESULTS
        # ====================================================================
        
        backup_to_load = grid_to_load + diesel_to_load
        
        # Calculate total load ACTUALLY served
        total_load_demand = np.sum(load)
        total_unmet = np.sum(unmet_load)
        total_load_served = total_load_demand - total_unmet
        self_sufficiency = (total_load_served / total_load_demand * 100) if total_load_demand > 0 else 0
        
        results = {
            'year': year,
            'pv_generation_kwh': np.sum(pv_generation),
            'load_demand_kwh': total_load_demand,
            'load_served_kwh': total_load_served,
            'unmet_load_kwh': total_unmet,
            'self_sufficiency_%': self_sufficiency,
            'pv_to_load_kwh': np.sum(pv_to_load),
            'pv_to_battery_kwh': np.sum(pv_to_battery),
            'battery_to_load_kwh': np.sum(battery_to_load),
            'backup_kwh': np.sum(backup_to_load),
            'grid_kwh': np.sum(grid_to_load),
            'diesel_kwh': np.sum(diesel_to_load),
            'curtailed_kwh': np.sum(curtailed),
            'generator_spillage_kwh': np.sum(generator_spillage),
            'renewable_fraction': (np.sum(pv_to_load) + np.sum(battery_to_load)) / total_load_served if total_load_served > 0 else 0,
            'soc_profile': np.array(soc_profile),
            'hourly_pv': pv_generation,
            'hourly_load': load,
            'hourly_backup': backup_to_load,
            'hourly_grid': grid_to_load,
            'hourly_diesel': diesel_to_load,
            'hourly_unmet': unmet_load,
            'hourly_generator_spillage': generator_spillage,
            'hourly_grid_to_battery': grid_to_battery,
            'hourly_diesel_to_battery': generator_to_battery,    
            'hourly_battery_discharge': battery_to_load,
            'hourly_battery_charge': pv_to_battery + generator_to_battery + grid_to_battery,
            'hourly_grid_tariff': grid_tariff_profile,
            'grid_to_battery_kwh': np.sum(grid_to_battery),
            'diesel_to_battery_kwh': np.sum(generator_to_battery)  
        }
        
        # ====================================================================
        # 6. CALCULATE BACKUP COSTS
        # ====================================================================
        if self.backup_type in ['Diesel', 'Grid+Diesel']:
            fuel_consumed = (np.sum(diesel_to_load) + np.sum(generator_to_battery)) * self.diesel_fuel_consumption
            results['diesel_fuel_liters'] = fuel_consumed
            results['diesel_cost_usd'] = fuel_consumed * self.diesel_fuel_price
        
        if self.backup_type in ['Grid', 'Grid+Diesel']:
            total_grid_consumption = grid_to_load + grid_to_battery
            grid_cost = np.sum(total_grid_consumption * grid_tariff_profile)
            
            results['grid_energy_kwh'] = np.sum(grid_to_load)
            results['grid_charging_kwh'] = np.sum(grid_to_battery)
            results['grid_total_kwh'] = np.sum(total_grid_consumption)
            results['grid_cost_usd'] = grid_cost
            
            peak_rate = self.peak_grid_tariff
            battery_discharge_value = np.sum(battery_to_load) * peak_rate
            charging_cost = np.sum(grid_to_battery * grid_tariff_profile)
            results['arbitrage_savings_usd'] = battery_discharge_value - charging_cost
        
        # ====================================================================
        # 7. ADD BATTERY HEALTH INFO
        # ====================================================================
        if self.bess is not None:
            results['battery_capacity_remaining'] = self.bess.capacity_kwh / self.bess.nominal_capacity_kwh
            results['battery_throughput_kwh'] = self.bess.cumulative_throughput_kwh
        
        return results

    def get_grid_tariff(self, hour_of_day, noon_grid_tariff, peak_grid_tariff, night_grid_tariff):
        
        if self.backup_type not in ['Grid', 'Grid+Diesel']:
            return 0
        
        if  (6 <= hour_of_day < 17) :
            return noon_grid_tariff
        
        elif (17 <= hour_of_day < 22) :    
            return peak_grid_tariff
        
        else:
            return night_grid_tariff

    def estimate_next_24h_pv_excess(self, current_hour_index, pv_generation):
        
        if self.pv is None or self.bess is None:
            return 0
        
        hours_to_check = min(24, 8760 - current_hour_index)
        
        load_ahead = self.load_profile['Load_kW'].values[
            current_hour_index : current_hour_index + hours_to_check
        ]
        
        pv_ahead = pv_generation[current_hour_index : current_hour_index + hours_to_check]
        
        potential_excess = np.maximum(0, pv_ahead - load_ahead)
        expected_pv_excess_kwh = np.sum(potential_excess)
        
        return expected_pv_excess_kwh
    
    def should_charge_from_grid(self, current_hour, current_tariff, current_hour_index, 
                                pv_generation, current_soc):
        
        if self.backup_type not in ['Grid', 'Grid+Diesel'] or self.bess is None:
            return False
        
        if current_soc > 0.9:
            return False
        
        expected_pv_excess = self.estimate_next_24h_pv_excess(
            current_hour_index, pv_generation
        )
        
        battery_capacity_available = (self.bess.soc_max - current_soc) * self.bess.capacity_kwh
        
        if expected_pv_excess >  battery_capacity_available:
            return False 
        
        effective_cost = current_tariff / self.bess.efficiency
        
        return effective_cost < self.peak_grid_tariff
    
    def simulate_multi_year(self, num_years, replacement_threshold):
        """Multi-year simulation with unmet load tracking"""
        if replacement_threshold is None:
            raise ValueError("HybridEnergySimulator.simulate_multi_year: 'replacement_threshold' is required and cannot be None")
        annual_results = []
        years_since_replacement = 0
                    
        for year in tqdm(range(num_years), desc="Simulating years", leave=False):
            battery_capacity_ratio = 1.0  # Default for no battery
            if self.bess is not None:
                battery_capacity_ratio = self.bess.update_capacity(years_since_replacement)
                if replacement_threshold > 0 and battery_capacity_ratio < replacement_threshold and year > 0:
                    self.bess.replace_battery(year)
                    battery_capacity_ratio = 1.0
                    years_since_replacement = 0      
                else:
                    years_since_replacement += 1
            
            results = self.simulate_year(year=year)
            results['battery_capacity_ratio'] = battery_capacity_ratio  # Store capacity ratio
            annual_results.append(results)
        
        # Create summary
        summary = {
            'annual_results': annual_results,
            'total_pv_generation_kwh': sum(r['pv_generation_kwh'] for r in annual_results),
            'total_load_demand_kwh': sum(r['load_demand_kwh'] for r in annual_results),
            'total_load_served_kwh': sum(r['load_served_kwh'] for r in annual_results),
            'total_unmet_load_kwh': sum(r['unmet_load_kwh'] for r in annual_results),
            'avg_self_sufficiency_%': np.mean([r['self_sufficiency_%'] for r in annual_results]),
            'total_backup_kwh': sum(r['backup_kwh'] for r in annual_results),
            'total_curtailed_kwh': sum(r['curtailed_kwh'] for r in annual_results),
            'total_generator_spillage_kwh': sum(r.get('generator_spillage_kwh', 0) for r in annual_results),
            'avg_renewable_fraction': np.mean([r['renewable_fraction'] for r in annual_results]),
            'total_grid_kwh': sum(r.get('grid_kwh', 0) for r in annual_results),
            'total_diesel_kwh': sum(r.get('diesel_kwh', 0) for r in annual_results),
        }
        
        if self.bess is not None:
            summary['battery_replacements'] = self.bess.replacement_count
            summary['replacement_years'] = self.bess.replacement_years
        
        if self.backup_type in ['Diesel', 'Grid+Diesel']:
            summary['total_diesel_cost'] = sum(r.get('diesel_cost_usd', 0) for r in annual_results)
        
        if self.backup_type in ['Grid', 'Grid+Diesel']:
            summary['total_grid_cost'] = sum(r.get('grid_cost_usd', 0) for r in annual_results)
        
        return summary

class FinancialAnalyzer:
    """Traditional LCOE calculation for renewable systems"""
    
    def __init__(self, discount_rate):
        if discount_rate is None:
            raise ValueError("FinancialAnalyzer: 'discount_rate' is required and cannot be None")
        self.discount_rate = discount_rate
    
    def calculate_lcoe(self, capex, annual_opex, energy_generated_annual, 
                       project_years, replacement_costs):
        """Calculate Levelized Cost of Energy"""
        if replacement_costs is None:
            raise ValueError("FinancialAnalyzer.calculate_lcoe: 'replacement_costs' is required. Pass {} if no replacements.")
        
        if isinstance(annual_opex, (int, float)):
            annual_opex = [annual_opex] * project_years
        
        npv_costs = capex
        
        for year in range(1, project_years + 1):
            year_cost = annual_opex[year - 1] if year - 1 < len(annual_opex) else annual_opex[-1]
            
            if year in replacement_costs:
                year_cost += replacement_costs[year]
            
            discount_factor = (1 + self.discount_rate) ** year
            npv_costs += year_cost / discount_factor
        
        npv_energy = 0
        for year in range(1, project_years + 1):
            year_idx = year - 1
            if year_idx < len(energy_generated_annual):
                energy = energy_generated_annual[year_idx]
            else:
                energy = 0
            
            discount_factor = (1 + self.discount_rate) ** year
            npv_energy += energy / discount_factor
        
        if npv_energy == 0:
            return float('inf')
        
        lcoe = npv_costs / npv_energy
        return lcoe
    
    def calculate_system_lcoe(self, simulation_results, capex_params, opex_params, 
                         project_years, backup_type,
                         diesel_escalation, grid_escalation,
                         include_grid_electricity_cost=True):
        """Calculate LCOE for complete hybrid system"""
        annual_results = simulation_results['annual_results']
        
        pv_useful_energy = [
            r.get('pv_to_load_kwh', 0) + r.get('pv_to_battery_kwh', 0) 
            for r in annual_results
        ]
        
        bess_discharge = [r.get('battery_to_load_kwh', 0) for r in annual_results]
        backup_energy = [r.get('backup_kwh', 0) for r in annual_results]
        total_load = [r.get('load_served_kwh', 0) for r in annual_results]
        
        renewable_useful_energy = pv_useful_energy
        
        # LCOE for renewable system
        renewable_capex = sum([
            capex_params.get('pv_module', 0),
            capex_params.get('mounting', 0),
            capex_params.get('dc_bos', 0),
            capex_params.get('other', 0),
            capex_params.get('scada', 0),
            capex_params.get('bess_battery', 0),
            capex_params.get('bess_pcs', 0),
            capex_params.get('bess_bos', 0)
        ])
        
        renewable_opex = (
            opex_params.get('pv_om', 0) +
            opex_params.get('bess_om', 0)
        )
        
        replacement_costs = {}
        if 'replacement_years' in simulation_results:
            for year in simulation_results['replacement_years']:
                replacement_costs[year] = (
                    capex_params.get('bess_battery', 0) +
                    capex_params.get('bess_pcs', 0) +
                    capex_params.get('bess_bos', 0)
                )
        
        if sum(renewable_useful_energy) > 0:
            lcoe_renewable = self.calculate_lcoe(
                renewable_capex, 
                renewable_opex, 
                renewable_useful_energy,
                project_years,
                replacement_costs
            )
        else:
            lcoe_renewable = 0
        
        # LCOE for backup system
        backup_capex = capex_params.get('diesel_system', 0) + capex_params.get('grid_system', 0)
        
        backup_opex_annual = []
        for year_idx, r in enumerate(annual_results):
            base_opex = opex_params.get('diesel_om', 0) + opex_params.get('grid_om', 0)
            
            if backup_type in ['Diesel', 'Grid+Diesel']:
                fuel_cost = r.get('diesel_cost_usd', 0)
                fuel_cost_escalated = fuel_cost * (1 + diesel_escalation) ** year_idx
            else:
                fuel_cost_escalated = 0
            
            if backup_type in ['Grid', 'Grid+Diesel']:
                grid_cost = r.get('grid_cost_usd', 0)
                # Only include grid electricity cost if flag is True
                # If OPEX_Grid = 0 (customer pays), this will be False
                if include_grid_electricity_cost:
                    grid_cost_escalated = grid_cost * (1 + grid_escalation) ** year_idx
                else:
                    grid_cost_escalated = 0  # Exclude grid electricity purchases from YOUR LCOE
            else:
                grid_cost_escalated = 0
            
            backup_opex_annual.append(base_opex + fuel_cost_escalated + grid_cost_escalated)
        
        if sum(backup_energy) > 0 and backup_capex > 0:
            lcoe_backup = self.calculate_lcoe(
                backup_capex,
                backup_opex_annual,
                backup_energy,
                project_years,
                replacement_costs={}
            )
        else:
            lcoe_backup = 0
        
        # Total system LCOE
        total_capex = sum(capex_params.values())
        
        total_opex_annual = [
            renewable_opex + backup_opex_annual[i]
            for i in range(project_years)
        ]
        
        lcoe_total = self.calculate_lcoe(
            total_capex,
            total_opex_annual,
            total_load,
            project_years,
            replacement_costs
        )
        
        # Component LCOEs
        pv_capex = sum([
            capex_params.get('pv_module', 0),
            capex_params.get('mounting', 0),
            capex_params.get('other', 0),
            capex_params.get('dc_bos', 0),
            capex_params.get('scada', 0)
        ])
        pv_opex = opex_params.get('pv_om', 0)
        
        if sum(pv_useful_energy) > 0 and pv_capex > 0:
            lcoe_pv = self.calculate_lcoe(
                pv_capex, 
                pv_opex, 
                pv_useful_energy,
                project_years,
                replacement_costs={}
            )
        else:
            lcoe_pv = 0
        
        bess_capex = sum([
            capex_params.get('bess_battery', 0),
            capex_params.get('bess_pcs', 0),
            capex_params.get('bess_bos', 0)
        ])
        bess_opex = opex_params.get('bess_om', 0)
        
        if sum(bess_discharge) > 0 and bess_capex > 0:
            lcoe_bess = self.calculate_lcoe(
                bess_capex, 
                bess_opex, 
                bess_discharge,
                project_years,
                replacement_costs
            )
        else:
            lcoe_bess = 0
        
        return {
            'lcoe_total': lcoe_total,
            'lcoe_renewable': lcoe_renewable,
            'lcoe_pv': lcoe_pv,
            'lcoe_bess': lcoe_bess,
            'lcoe_backup': lcoe_backup,
            'total_capex': total_capex,
            'renewable_fraction': simulation_results.get('avg_renewable_fraction', 0),
            'battery_replacements': simulation_results.get('battery_replacements', 0)
        }

def generate_comprehensive_excel(optimal_config, simulation_results, power_kw, load_df, simulation_years,
                                 capex_params, opex_params, discount_rate,
                                 all_configs_df=None, backup_type='None', latitude=0, longitude=0,
                                 diesel_escalation=0, grid_escalation=0):
    
    excel_buffer = BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        
        pv_cap_kw = optimal_config.get('PV_kWp', 0)
        bess_kwh = optimal_config.get('BESS_kWh', 0)
        bess_efficiency = optimal_config.get('Round_Trip_Efficiency', 85)
        
        annual_load = simulation_results.get('total_load_demand_kwh', 0) / simulation_years / 1000
        annual_pv = simulation_results.get('total_pv_generation_kwh', 0) / simulation_years / 1000
        
        # SHEET 1: EXECUTIVE SUMMARY
        exec_summary = {
            'Metric': [
                '=== PROJECT CONFIGURATION ===', 'Configuration Name', 'PV Capacity (kWp)', 'BESS Capacity (kWh)',
                'BESS Power (kW)', 'BESS Duration (hours)', '',
                '=== FINANCIAL SUMMARY ===', 'Total CAPEX ($)', 'LCOE PV ($/kWh)', 'LCOE BESS ($/kWh)',
                'LCOE Renewable ($/kWh)', 'LCOE Total with Backup ($/kWh)', '',
                '=== ENERGY SUMMARY ===', 'Total Load (MWh/year)', 'PV Generation (MWh/year)', 'Renewable Fraction (%)', '',
                '=== PERFORMANCE METRICS ===', 'PV Capacity Factor (DC) (%)',
                'BESS Round-trip Efficiency (%)', 'BESS Utilization (%)', '',
                '=== ENERGY FLOWS ===', 'BESS Charge (MWh/year)', 'BESS Discharge (MWh/year)',
                'Backup Used (MWh/year)', 'Curtailed Energy (MWh/year)', 'Curtailment Rate (%)', '',
                '=== BATTERY DETAILS ===', 'Battery Replacements', 'Replacement Years',
                'Final Battery Capacity (%)', 'Total Battery Throughput (MWh)', '',
                '=== BACKUP SYSTEM ===', 'Backup Type', 'Max Backup Power (kW)',
                'Backup Energy Cost ($/year)', 'Backup Capacity Factor (%)', '',
                '=== PROJECT DETAILS ===', 'Simulation Years', 'Discount Rate (%)',
                'Location', 'Report Generated'
            ],
            'Value': []
        }
        
        if pv_cap_kw > 0:
            pv_cf_ac = (annual_pv * 1000) / (pv_cap_kw * 8760) * 100
        else:
            pv_cf_ac = 0
        
        if bess_kwh > 0:
            annual_bess_charge = sum(r.get('grid_to_battery_kwh', 0) + r.get('diesel_to_battery_kwh', 0) + r.get('pv_to_battery_kwh', 0) for r in simulation_results.get('annual_results', [])) / simulation_years / 1000
            annual_bess_discharge = sum(r.get('battery_to_load_kwh', 0) for r in simulation_results.get('annual_results', [])) / simulation_years / 1000
            bess_utilization = (annual_bess_discharge * 1000) / (bess_kwh * 365) * 100
            final_bess_capacity = simulation_results['annual_results'][-1].get('battery_capacity_remaining', 1.0) * 100
            total_throughput = simulation_results['annual_results'][-1].get('battery_throughput_kwh', 0) / 1000
        else:
            annual_bess_charge = annual_bess_discharge = bess_utilization = final_bess_capacity = total_throughput = 0
        
        annual_backup = simulation_results.get('total_backup_kwh', 0) / simulation_years / 1000
        max_backup_kw = max([max(r.get('hourly_backup', [0])) for r in simulation_results.get('annual_results', [])]) if simulation_results.get('annual_results') else 0
        backup_cf = (annual_backup * 1000) / (max_backup_kw * 8760) * 100 if max_backup_kw > 0 else 0
        
        annual_curtailed = simulation_results.get('total_curtailed_kwh', 0) / simulation_years / 1000
        curtailment_rate = (annual_curtailed / annual_pv * 100) if annual_pv > 0 else 0
        
        total_capex = optimal_config.get('Total_CAPEX_$', 0)
        
        exec_summary['Value'] = [
            '', f"PV{pv_cap_kw:.0f}_B{bess_kwh:.0f}", f"{pv_cap_kw:.1f}", f"{bess_kwh:.1f}",
            f"{power_kw:.1f}", "2.0", '', '',
            f"${total_capex:,.0f}", f"${optimal_config.get('LCOE_PV_$/kWh', 0):.4f}",
            f"${optimal_config.get('LCOE_BESS_$/kWh', 0):.4f}", 
            f"${optimal_config.get('LCOE_Renewable_$/kWh', 0):.4f}",
            f"${optimal_config.get('LCOE_Total_$/kWh', 0):.4f}",
            '', '',
            f"{annual_load:.1f}", f"{annual_pv:.1f}", f"{optimal_config.get('Renewable_Fraction_%', 0):.1f}", '', '',
            f"{pv_cf_ac:.2f}", f"{bess_efficiency:.1f}", f"{bess_utilization:.1f}", '', '',
            f"{annual_bess_charge:.1f}", f"{annual_bess_discharge:.1f}", f"{annual_backup:.1f}",
            f"{annual_curtailed:.1f}", f"{curtailment_rate:.1f}", '', '',
            f"{simulation_results.get('battery_replacements', 0)}",
            ', '.join(map(str, simulation_results.get('replacement_years', []))) if simulation_results.get('replacement_years') else 'None',
            f"{final_bess_capacity:.1f}", f"{total_throughput:.1f}", '', '',
            backup_type, f"{max_backup_kw:.1f}", f"${(simulation_results.get('total_diesel_cost', 0) + simulation_results.get('total_grid_cost', 0)) / simulation_years:,.0f}",
            f"{backup_cf:.1f}", '', '', f"{simulation_years}", f"{discount_rate * 100:.1f}", f"Lat: {latitude}, Lon: {longitude}",
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        pd.DataFrame(exec_summary).to_excel(writer, sheet_name='Executive_Summary', index=False)
        
        # SHEET 2: CAPEX BREAKDOWN
        capex_breakdown = []
        if pv_cap_kw > 0:
            for comp, key in [('PV Modules', 'pv_module'), ('Mounting Structure', 'mounting'), 
                            ('DC BOS', 'dc_bos'), ('SCADA System', 'scada'), ('Other', 'other')]:
                capex_breakdown.append({
                    'Category': 'PV System', 'Component': comp, 
                    'Unit': 'lump sum' if comp == 'SCADA System' else '$/kW',
                    'Quantity': 1 if comp == 'SCADA System' else pv_cap_kw,
                    'Unit_Cost': capex_params.get(key, 0) if comp == 'SCADA System' else capex_params.get(key, 0) / pv_cap_kw,
                    'Total_Cost': capex_params.get(key, 0)
                })
        
        if bess_kwh > 0:
            for comp, key in [('Battery Cells', 'bess_battery'), ('Power Conversion System', 'bess_pcs'), 
                            ('BOS & Integration', 'bess_bos')]:
                capex_breakdown.append({
                    'Category': 'BESS', 'Component': comp, 'Unit': '$/kWh', 'Quantity': bess_kwh,
                    'Unit_Cost': capex_params.get(key, 0) / bess_kwh, 'Total_Cost': capex_params.get(key, 0)
                })
        
        if 'diesel_system' in capex_params and capex_params['diesel_system'] > 0:
            capex_breakdown.append({'Category': 'Backup', 'Component': 'Diesel Generator', 'Unit': 'lump sum', 
                                  'Quantity': 1, 'Unit_Cost': capex_params['diesel_system'], 'Total_Cost': capex_params['diesel_system']})
        if 'grid_system' in capex_params and capex_params['grid_system'] > 0:
            capex_breakdown.append({'Category': 'Backup', 'Component': 'Grid Connection', 'Unit': 'lump sum',
                                  'Quantity': 1, 'Unit_Cost': capex_params['grid_system'], 'Total_Cost': capex_params['grid_system']})
        
        capex_df = pd.DataFrame(capex_breakdown)
        total_row = pd.DataFrame([{'Category': 'TOTAL CAPEX', 'Component': '', 'Unit': '', 
                                  'Quantity': '', 'Unit_Cost': '', 'Total_Cost': capex_df['Total_Cost'].sum()}])
        capex_df = pd.concat([capex_df, total_row], ignore_index=True)
        capex_df.to_excel(writer, sheet_name='CAPEX_Breakdown', index=False)
        
        # SHEET 3: HOURLY YEAR 1
        if simulation_results and 'annual_results' in simulation_results:
            year_0 = simulation_results['annual_results'][0]
            hourly_data = pd.DataFrame({
                'Hour': range(1, 8761),
                'DateTime': pd.date_range(start='2025-01-01 00:00:00', periods=8760, freq='h'),
                'Load_kW': load_df['Load_kW'].values[:8760],
                'PV_Generation_kW': year_0.get('hourly_pv', np.zeros(8760))[:8760],
            })
            hourly_data['Total_Renewable_kW'] = hourly_data['PV_Generation_kW']
            hourly_data['BESS_Charge_kW'] = year_0.get('hourly_battery_charge', np.zeros(8760))[:8760]
            hourly_data['BESS_Discharge_kW'] = year_0.get('hourly_battery_discharge', np.zeros(8760))[:8760]
            hourly_data['Unmet_Load_kW'] = year_0.get('hourly_unmet', np.zeros(8760))[:8760]
            
            if bess_kwh > 0:
                hourly_data['BESS_SOC_%'] = year_0.get('soc_profile', np.zeros(8760))[:8760] * 100
                hourly_data['BESS_Energy_kWh'] = year_0.get('soc_profile', np.zeros(8760))[:8760] * bess_kwh
            else:
                hourly_data['BESS_SOC_%'] = hourly_data['BESS_Energy_kWh'] = 0
            
            hourly_data['Grid_to_Load_kW'] = year_0.get('hourly_grid', np.zeros(8760))[:8760]
            hourly_data['Grid_to_Battery_kW'] = year_0.get('hourly_grid_to_battery', np.zeros(8760))[:8760]
            hourly_data['Grid_Total_Generation_kW'] = hourly_data['Grid_to_Load_kW'] + hourly_data['Grid_to_Battery_kW']
            
            hourly_data['Diesel_to_Load_kW'] = year_0.get('hourly_diesel', np.zeros(8760))[:8760]
            hourly_data['Diesel_to_Battery_kW'] = year_0.get('hourly_diesel_to_battery', np.zeros(8760))[:8760]
            hourly_data['Diesel_Total_Generation_kW'] = hourly_data['Diesel_to_Load_kW'] + hourly_data['Diesel_to_Battery_kW']
            hourly_data['Generator_Spillage_kW'] = year_0.get('hourly_generator_spillage', np.zeros(8760))[:8760]
            
            hourly_data['Total_Backup_to_Load_kW'] = hourly_data['Grid_to_Load_kW'] + hourly_data['Diesel_to_Load_kW']
            hourly_data['Total_Backup_Generation_kW'] = hourly_data['Grid_Total_Generation_kW'] + hourly_data['Diesel_Total_Generation_kW']
            hourly_data['Curtailed_kW'] = np.maximum(0, hourly_data['Total_Renewable_kW'] - hourly_data['Load_kW'] - hourly_data['BESS_Charge_kW'])
            hourly_data['Renewable_Fraction_%'] = np.where(
                hourly_data['Load_kW'] > 0,
                ((hourly_data['PV_Generation_kW'] + hourly_data['BESS_Discharge_kW']) / hourly_data['Load_kW'] * 100).clip(0, 100), 0
            )
            
            # ADD STATUS COLUMN
            hourly_data['Status'] = np.where(
                hourly_data['Unmet_Load_kW'] > 0.01,  # Threshold to avoid floating point errors
                'DEMAND NOT MET ⚠️',
                'OK'
            )
            
            hourly_data.to_excel(writer, sheet_name='Hourly_Year_1', index=False)
            
            # APPLY CONDITIONAL FORMATTING TO HIGHLIGHT UNMET DEMAND
            worksheet = writer.sheets['Hourly_Year_1']
            from openpyxl.styles import PatternFill
            
            # Find the column index for Status and Unmet_Load_kW
            status_col = hourly_data.columns.get_loc('Status') + 1  # +1 because Excel is 1-indexed
            unmet_col = hourly_data.columns.get_loc('Unmet_Load_kW') + 1
            
            # Red fill for unmet demand rows
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            
            # Apply formatting to rows with unmet demand
            for row_idx in range(2, len(hourly_data) + 2):  # Start from row 2 (after header)
                if hourly_data.iloc[row_idx - 2]['Unmet_Load_kW'] > 0.01:
                    # Highlight the entire row or just key columns
                    worksheet.cell(row=row_idx, column=status_col).fill = red_fill
                    worksheet.cell(row=row_idx, column=unmet_col).fill = red_fill
        
        # SHEET 4: ANNUAL SUMMARY
        annual_summary = []
        for year_idx, year_result in enumerate(simulation_results.get('annual_results', [])):
            annual_summary.append({
                'Year': year_idx + 1,
                'Load_MWh': year_result.get('load_served_kwh', 0) / 1000,
                'PV_Generation_MWh': year_result.get('pv_generation_kwh', 0) / 1000,
                'Total_Renewable_MWh': year_result.get('pv_generation_kwh', 0) / 1000,
                'BESS_Charge_MWh': (year_result.get('pv_to_battery_kwh', 0) + year_result.get('grid_to_battery_kwh', 0) + year_result.get('diesel_to_battery_kwh', 0)) / 1000,
                'BESS_Discharge_MWh': year_result.get('battery_to_load_kwh', 0) / 1000,
                'Backup_MWh': year_result.get('backup_kwh', 0) / 1000,
                'Curtailed_MWh': year_result.get('curtailed_kwh', 0) / 1000,
                'Generator_Spillage_MWh': year_result.get('generator_spillage_kwh', 0) / 1000,
                'Renewable_Fraction_%': year_result.get('renewable_fraction', 0) * 100,
                'Battery_Capacity_%': year_result.get('battery_capacity_remaining', 1.0) * 100,
                'Battery_Throughput_MWh': year_result.get('battery_throughput_kwh', 0) / 1000
            })
        pd.DataFrame(annual_summary).to_excel(writer, sheet_name='Annual_Summary', index=False)
        
        # SHEET 5: MONTHLY STATISTICS
        if simulation_results and 'annual_results' in simulation_results:
            year_0 = simulation_results['annual_results'][0]
            dt_index = pd.date_range(start='2025-01-01 00:00:00', periods=8760, freq='h')
            hourly_with_date = pd.DataFrame({
                'datetime': dt_index,
                'load': load_df['Load_kW'].values[:8760],
                'pv': year_0.get('hourly_pv', np.zeros(8760))[:8760],
                'backup': year_0.get('hourly_backup', np.zeros(8760))[:8760],
                'bess_charge': year_0.get('hourly_battery_charge', np.zeros(8760))[:8760],
                'bess_discharge': year_0.get('hourly_battery_discharge', np.zeros(8760))[:8760],
                'generator_spillage': year_0.get('hourly_generator_spillage', np.zeros(8760))[:8760],
            })
            hourly_with_date['curtailed'] = np.maximum(0, hourly_with_date['pv'] - hourly_with_date['load'] - hourly_with_date['bess_charge'])
            hourly_with_date['renewable_delivered'] = hourly_with_date['pv'] - hourly_with_date['curtailed'] - hourly_with_date['bess_charge'] + hourly_with_date['bess_discharge']
            hourly_with_date['month'] = hourly_with_date['datetime'].dt.month
            
            numeric_cols = ['load', 'pv', 'backup', 'bess_charge', 'bess_discharge', 'curtailed', 'renewable_delivered', 'generator_spillage']
            monthly_grouped = hourly_with_date.groupby('month')[numeric_cols].sum()
            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            monthly_data = []
            
            for month in range(1, 13):
                monthly_data.append({
                    'Month': month, 'Month_Name': month_names[month - 1],
                    'Load_MWh': monthly_grouped.loc[month, 'load'] / 1000,
                    'PV_Generation_MWh': monthly_grouped.loc[month, 'pv'] / 1000,
                    'Total_Renewable_Gen_MWh': monthly_grouped.loc[month, 'pv'] / 1000,
                    'BESS_Charge_MWh': monthly_grouped.loc[month, 'bess_charge'] / 1000,
                    'BESS_Discharge_MWh': monthly_grouped.loc[month, 'bess_discharge'] / 1000,
                    'Renewable_Delivered_MWh': monthly_grouped.loc[month, 'renewable_delivered'] / 1000,
                    'Curtailed_MWh': monthly_grouped.loc[month, 'curtailed'] / 1000,
                    'Generator_Spillage_MWh': monthly_grouped.loc[month, 'generator_spillage'] / 1000,
                    'Backup_MWh': monthly_grouped.loc[month, 'backup'] / 1000,
                    'Renewable_Fraction_%': (monthly_grouped.loc[month, 'renewable_delivered'] / monthly_grouped.loc[month, 'load'] * 100) if monthly_grouped.loc[month, 'load'] > 0 else 0,
                    'Curtailment_Rate_%': (monthly_grouped.loc[month, 'curtailed'] / monthly_grouped.loc[month, 'pv'] * 100) if monthly_grouped.loc[month, 'pv'] > 0 else 0
                })
            
            monthly_df = pd.DataFrame(monthly_data)
            annual_totals = {'Month': 13, 'Month_Name': 'ANNUAL',
                           'Load_MWh': monthly_df['Load_MWh'].sum(),
                           'PV_Generation_MWh': monthly_df['PV_Generation_MWh'].sum(),
                           'Total_Renewable_Gen_MWh': monthly_df['Total_Renewable_Gen_MWh'].sum(),
                           'BESS_Charge_MWh': monthly_df['BESS_Charge_MWh'].sum(),
                           'BESS_Discharge_MWh': monthly_df['BESS_Discharge_MWh'].sum(),
                           'Renewable_Delivered_MWh': monthly_df['Renewable_Delivered_MWh'].sum(),
                           'Curtailed_MWh': monthly_df['Curtailed_MWh'].sum(),
                           'Generator_Spillage_MWh': monthly_df['Generator_Spillage_MWh'].sum(),
                           'Backup_MWh': monthly_df['Backup_MWh'].sum(),
                           'Renewable_Fraction_%': (monthly_df['Renewable_Delivered_MWh'].sum() / monthly_df['Load_MWh'].sum() * 100) if monthly_df['Load_MWh'].sum() > 0 else 0,
                           'Curtailment_Rate_%': (monthly_df['Curtailed_MWh'].sum() / monthly_df['Total_Renewable_Gen_MWh'].sum() * 100) if monthly_df['Total_Renewable_Gen_MWh'].sum() > 0 else 0}
            monthly_df = pd.concat([monthly_df, pd.DataFrame([annual_totals])], ignore_index=True)
            monthly_df.to_excel(writer, sheet_name='Monthly_Statistics', index=False)

        
        # SHEET 6: CONFIGURATION
        config_data = [{'Parameter': k, 'Value': str(v)} for k, v in optimal_config.items() if k != 'sim_results']
        pd.DataFrame(config_data).to_excel(writer, sheet_name='Configuration', index=False)
        
        # SHEET 7: ALL CONFIGURATIONS (if provided)
        if all_configs_df is not None and len(all_configs_df) > 0:
            all_configs_df.to_excel(writer, sheet_name='All_Configurations', index=False)
    
    excel_buffer.seek(0)
    return excel_buffer.getvalue()
